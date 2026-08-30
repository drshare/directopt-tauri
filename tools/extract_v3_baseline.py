#!/usr/bin/env python3
"""从「绿电直连新能源优化配置 V3.0」真实输出生成对拍基准夹具。

用法：
    python tools/extract_v3_baseline.py <真实输出.xlsx> <输入模板.xlsx>

输入：
    1. 真实输出 xlsx（V3.0「下载计算结果」导出）
       - curveldzl3 Sheet：全年 8760h 输入曲线（8 列）
       - output Sheet   ：输出指标（序号 1~34）+ 逐时电量平衡表（8760 行 × 13 列）
    2. 输入模板 xlsx（inputtemplate_ldzl_3.0.xlsx）——用于还原计算参数

产出（写入 src-tauri/src/compute/）：
    v3_fixture.json  参数 + 输入曲线（供 Rust 端重建计算）
    v3_expected.json 输出指标（序号 1~34，公式无缓存值的项由逐时数据推导）
    v3_hourly.json   逐时电量平衡表（8760 行 × 13 列）
"""

import json
import sys
from pathlib import Path

import openpyxl

HOURS = 8760

# 输入模板「项目」→ ComputeParams 字段（与 src/lib/fileParsers.ts 保持一致）
TEMPLATE_MAP = {
    "储能充放电深度": "dod",
    "电池充放电倍率": "rate",
    "储能初始电量": "initialSoc",
    "储能系统充电效率": "chargeEff",
    "储能系统放电效率": "dischargeEff",
    "接入公共电网容量（最大下网功率）": "gridCapacity",
    "平均负荷率": "avgLoadRate",
    "自发自用占总可用发电量比例下限": "selfUseGenMin",
    "自发自用占总用电量比例下限": "selfUseLoadMin",
    "余电上网比例上限": "feedLimit",
    "余电最大上网功率": "feedPower",
    "弃电率上限": "curtailLimit",
    "风电系统单位投资": "windInvest",
    "光伏系统单位投资": "pvInvest",
    "储能系统单位投资": "essInvest",
    "年运维费用占比": "opexRatio",
    "人员工资": "salary",
    "定员人数": "staffCount",
    "折现率": "discountRate",
    "评价周期": "evalPeriod",
    "其他固定投资": "otherInvest",
    "电池更换单价": "batteryReplaceUnit",
    "电池更换比例": "batteryReplaceRatio",
    "电池更换时间": "batteryReplaceYear",
    "选定风电规模起始值": "windStart",
    "选定风电规模结束值": "windEnd",
    "选定光伏规模起始值": "pvStart",
    "选定光伏规模结束值": "pvEnd",
    "选定储能容量起始值": "essStart",
    "选定储能容量结束值": "essEnd",
    "总评估次数": "nIter",
    "初始随机采样点数": "nInit",
}

# 逐时表列顺序（与 output Sheet C~P 列一致）
HOURLY_COLUMNS = [
    "时间",
    "风电发电量（kWh）",
    "光伏发电量（kWh）",
    "新能源理论发电量（kWh）",
    "用户负荷（kWh）",
    "该小时段新能源实际发电量（kWh）",
    "该小时段储能充电量（交流侧）（kWh）",
    "该小时段储能实际充电量（直流侧）（kWh）",
    "该小时段弃风弃光电量(kWh)",
    "该小时段储能放电量（直流侧）（kWh）",
    "该小时段储能对外供电量（交流测）（kWh）",
    "该小时段下网电量(kWh)",
    "该小时段余电上网电量(kWh)",
    "储能可用电量（直流侧）（kWh）",
]


def read_template(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    li, vi = header.index("项目"), header.index("数值")
    out = {}
    for r in rows[1:]:
        if li >= len(r) or r[li] is None:
            continue
        label = str(r[li]).strip()
        if label in TEMPLATE_MAP and vi < len(r) and r[vi] is not None:
            out[TEMPLATE_MAP[label]] = float(r[vi])
    missing = [k for k in ("windStart", "nIter") if k not in out]
    if missing:
        raise SystemExit(f"模板缺少关键参数：{missing}")
    return out


def read_curves(path: Path) -> dict:
    """读取 curveldzl3 Sheet 的 8760h 输入曲线（前 8 列）。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "curveldzl3" not in wb.sheetnames:
        raise SystemExit(f"{path} 缺少 curveldzl3 工作表")
    ws = wb["curveldzl3"]
    # 列布局：0=时间，1=用电负荷，2=风电标幺，3=光伏标幺，
    #        4=电量电价，5=线损费，6=电度输配电价，7=系统运行费，8=政府性基金及附加
    col_start = 1
    cols: list[list[float]] = [[] for _ in range(8)]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 表头
        if len(cols[0]) >= HOURS:
            break
        for j in range(8):
            v = r[col_start + j] if col_start + j < len(r) else 0.0
            cols[j].append(float(v) if isinstance(v, (int, float)) else 0.0)
    for j, c in enumerate(cols):
        if len(c) != HOURS:
            raise SystemExit(f"曲线第 {j + 1} 列长度为 {len(c)}，应为 {HOURS}")
    return {
        "load": cols[0],
        "windPu": cols[1],
        "pvPu": cols[2],
        "price": cols[3],
        "lossFee": cols[4],
        "tduFee": cols[5],
        "systemFee": cols[6],
        "fundFee": cols[7],
    }


def read_output(path: Path):
    """返回（指标 dict, 逐时 8760×13）。指标取 F 列，缺失时回退 G 列（备注中的计算值）。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["output"]
    rows = list(ws.iter_rows(values_only=True))

    metrics: dict[str, float | None] = {}
    for r in rows:
        if len(r) < 4:
            continue
        if r[1] is None or not str(r[1]).strip().isdigit():
            continue
        idx = int(str(r[1]).strip())
        if not (1 <= idx <= 34):
            continue
        label = str(r[2]).strip() if r[2] is not None else ""
        if not label:
            continue
        val = None
        for col in (5, 6):  # F=数据, G=备注（部分项存放计算值）
            if col < len(r) and isinstance(r[col], (int, float)):
                val = float(r[col])
                break
        metrics[label] = val

    # 逐时表：定位表头行（含「序号」「时间」）
    head_i = None
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if "序号" in cells and "时间" in cells:
            head_i = i
            break
    if head_i is None:
        raise SystemExit("output Sheet 中未找到逐时表表头")
    hourly = []
    for r in rows[head_i + 1 :]:
        if len(hourly) >= HOURS:
            break
        if r[1] is None or not str(r[1]).strip().isdigit():
            continue
        # 列布局：0=空 1=序号 2=时间 3..15=13 列数值 16=备注
        row = []
        for j in range(3, 16):  # 风电 … 储能可用电量
            v = r[j] if j < len(r) else 0.0
            row.append(float(v) if isinstance(v, (int, float)) else 0.0)
        hourly.append(row)
    if len(hourly) != HOURS:
        raise SystemExit(f"逐时表行数为 {len(hourly)}，应为 {HOURS}")
    return metrics, hourly


def derive_missing(metrics: dict, hourly: list[list[float]]) -> dict:
    """公式无缓存值的电量类指标由逐时数据推导（独立复算，不依赖对方公式）。"""
    col = {name: i for i, name in enumerate(HOURLY_COLUMNS[1:])}
    s = lambda name: sum(row[col[name]] for row in hourly) / 1000.0  # kWh → MWh

    wind, pv = s("风电发电量（kWh）"), s("光伏发电量（kWh）")
    theory = s("新能源理论发电量（kWh）")
    actual = s("该小时段新能源实际发电量（kWh）")
    curtailed = s("该小时段弃风弃光电量(kWh)")
    charge_ac = s("该小时段储能充电量（交流侧）（kWh）")
    charge_dc = s("该小时段储能实际充电量（直流侧）（kWh）")
    discharge_dc = s("该小时段储能放电量（直流侧）（kWh）")
    discharge_ac = s("该小时段储能对外供电量（交流测）（kWh）")
    grid = s("该小时段下网电量(kWh)")
    feed = s("该小时段余电上网电量(kWh)")
    load = s("用户负荷（kWh）")
    end_soc = hourly[-1][col["储能可用电量（直流侧）（kWh）"]] / 1000.0

    derived = {
        "全年风电最大发电量": wind,
        "全年光伏最大发电量": pv,
        "全年新能源最大发电量": theory,
        "全年弃风弃光电量": curtailed,
        "全年新能源实际发电量": actual,
        "全年储能充电量（交流侧）": charge_ac,
        "全年储能实际充电量（直流侧）": charge_dc,
        "全年储能实际放电量（直流侧）": discharge_dc,
        "全年储能供电量（交流侧）": discharge_ac,
        "全年下网电量": grid,
        "全年负荷用电量": load,
        "下网电量占比": grid / load * 100.0 if load else None,
        "绿电比": (1.0 - grid / load) * 100.0 if load else None,
        "储能年末剩余电量": end_soc,
    }
    for k, v in derived.items():
        if metrics.get(k) is None:
            metrics[k] = v
    return metrics


def main() -> None:
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    out_xlsx, tpl_xlsx = Path(sys.argv[1]), Path(sys.argv[2])
    dest = Path(__file__).resolve().parent.parent / "src-tauri/src/compute"

    tpl = read_template(tpl_xlsx)
    curves = read_curves(out_xlsx)
    metrics, hourly = read_output(out_xlsx)
    metrics = derive_missing(metrics, hourly)

    # 网电电价由「年电网购电成本 / 全年下网电量」推导（对方公式为 =F53/F43*10）
    if metrics.get("评价周期内平均网电电价") is None:
        buy = metrics.get("年电网购电成本")
        grid = metrics.get("全年下网电量")
        if buy and grid:
            metrics["评价周期内平均网电电价"] = buy / grid * 10.0

    params = {
        "tech": {
            "dod": tpl["dod"],
            "rate": tpl["rate"],
            "initialSoc": tpl["initialSoc"],
            "chargeEff": tpl["chargeEff"],
            "dischargeEff": tpl["dischargeEff"],
            "gridCapacity": tpl["gridCapacity"],
            "avgLoadRate": tpl["avgLoadRate"],
            "selfUseGenMin": tpl["selfUseGenMin"],
            "selfUseLoadMin": tpl["selfUseLoadMin"],
            "feedLimit": tpl["feedLimit"],
            "feedPower": tpl["feedPower"],
            "curtailLimit": tpl["curtailLimit"],
        },
        "econ": {
            "windInvest": tpl["windInvest"],
            "pvInvest": tpl["pvInvest"],
            "essInvest": tpl["essInvest"],
            "opexRatio": tpl["opexRatio"],
            "salary": tpl["salary"],
            "staffCount": tpl["staffCount"],
            "discountRate": tpl["discountRate"],
            "evalPeriod": tpl["evalPeriod"],
            "otherInvest": tpl["otherInvest"],
            "batteryReplaceUnit": tpl["batteryReplaceUnit"],
            "batteryReplaceRatio": tpl["batteryReplaceRatio"],
            "batteryReplaceYear": tpl["batteryReplaceYear"],
        },
        "algorithm": "bo",
        "bo": {"nIter": int(tpl["nIter"]), "nInit": int(tpl["nInit"])},
        "ga": {
            "generations": 40,
            "crossoverRate": 0.5,
            "mutationRate": 0.3,
            "populationSize": 100,
        },
        "range": {
            "windStart": tpl["windStart"],
            "windEnd": tpl["windEnd"],
            "pvStart": tpl["pvStart"],
            "pvEnd": tpl["pvEnd"],
            "essStart": tpl["essStart"],
            "essEnd": tpl["essEnd"],
        },
        "scheme": "scheme1",
        "objective": "composite",
        "chargePeriods": [],
        "dischargePeriods": [],
    }

    dest.mkdir(parents=True, exist_ok=True)
    (dest / "v3_fixture.json").write_text(
        json.dumps({"params": params, "curves": curves}, ensure_ascii=False),
        encoding="utf-8",
    )
    (dest / "v3_expected.json").write_text(
        json.dumps(metrics, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    (dest / "v3_hourly.json").write_text(
        json.dumps({"columns": HOURLY_COLUMNS[1:], "rows": hourly}, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"已生成夹具 → {dest}")
    print(f"  参数：范围 {tpl['windStart']}~{tpl['windEnd']} MW，BO({int(tpl['nIter'])}/{int(tpl['nInit'])})")
    print(f"  逐时：{len(hourly)} 行 × {len(hourly[0])} 列")
    print("  关键指标：")
    for k in (
        "最优风电规模",
        "最优光伏规模",
        "最优储能容量",
        "初投资",
        "年运行成本",
        "评价周期内总成本",
        "评价周期内平均综合电价",
        "评价周期内平均绿电电价",
        "弃电率" if "弃电率" in metrics else "全年弃风弃光电量",
    ):
        print(f"    {k}: {metrics.get(k)}")


if __name__ == "__main__":
    main()
