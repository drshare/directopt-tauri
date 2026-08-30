import openpyxl
import sys

for fname in ['docs/测试/inputtemplate_ldzl_3.0.xlsx', 'docs/测试/curvetemplate_ldzl_3.0.xlsx']:
    print('='*80)
    print(f'FILE: {fname}')
    print('='*80)
    wb = openpyxl.load_workbook(fname, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n--- Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---')
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(f'{cell.coordinate}={v}')
            if vals:
                print('  ' + ' | '.join(vals))
